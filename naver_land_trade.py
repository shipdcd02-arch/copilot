"""
네이버 부동산 아파트 데이터 수집
  · SQLite DB 누적 저장 (listings, listing_images, trades, daily_summary)
  · 신규 매물 상세정보 + 이미지 다운로드
  · xlsm 내보내기 – 날짜별 매물 조회 VBA 포함

사용법:
  py naver_land_trade.py                              # 기본값 (거제경남아너스빌)
  py naver_land_trade.py --no 12345                   # 단지번호만 지정 (이름 자동조회)
  py naver_land_trade.py --no 12345 --name 아파트명   # 이름 직접 지정
  py naver_land_trade.py --no 12345 --lawd 11110      # LAWD_CD도 직접 지정

단지번호 찾기: fin.land.naver.com 에서 아파트 검색 → URL의 숫자 (예: /complexes/12345)

[실거래가 조회]
  data.go.kr → "아파트매매 실거래가 상세 자료" 서비스키를 MOLIT_API_KEY에 입력
"""

import sys
import os
import json
import sqlite3
import requests
import time
import argparse
import xml.etree.ElementTree as ET
from datetime import datetime, date
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

# ── 고정 설정 ──────────────────────────────────────────────────────────────
MOLIT_MONTHS       = 6
FETCH_DETAIL       = True
DOWNLOAD_IMAGES    = True
MAX_DETAIL_PER_RUN = 30
MAX_SCRAPE_PER_RUN = 5
SCRAPE_DELAY_SEC   = (5, 12)

NAVER_COOKIES = {
    # 'NID_AUT': '...',
    # 'NID_SES': '...',
}
MOLIT_API_KEY = 'fd3f2f079d8b41080508a68e792161b95f0aac509e72efa527dfeba483c03075'
# ──────────────────────────────────────────────────────────────────────────

# ── 커맨드라인 인수 파싱 ───────────────────────────────────────────────────
_DEFAULT_COMPLEX = {
    # (단지번호): (이름, LAWD_CD)  ← 자주 쓰는 단지 미리 등록
    '105099': ('거제경남아너스빌', '48310'),
}

def _parse_args():
    p = argparse.ArgumentParser(add_help=False)
    p.add_argument('--no',   default='105099', help='fin.land 단지번호')
    p.add_argument('--name', default='',       help='아파트 이름 (생략 시 API 자동조회)')
    p.add_argument('--lawd', default='',       help='법정동 코드 앞 5자리 (생략 시 자동추출)')
    args, _ = p.parse_known_args()
    # 미리 등록된 단지면 name/lawd 기본값 적용
    if args.no in _DEFAULT_COMPLEX:
        d_name, d_lawd = _DEFAULT_COMPLEX[args.no]
        if not args.name: args.name = d_name
        if not args.lawd: args.lawd = d_lawd
    return args

_ARGS = _parse_args()

def _resolve_complex_info(no: str, name: str, lawd: str):
    """단지번호로 m.land API에서 이름/법정동코드 자동조회.
    429 등 실패 시 제공된 값 또는 기본값 사용.
    """
    if name and lawd:
        return name, lawd
    try:
        # m.land 목록 API에서 단지 정보 추출 (rate limit 없음)
        r = requests.get(
            'https://m.land.naver.com/complex/getComplexArticleList',
            params={'hscpNo': no, 'tradTpCd': 'A1', 'order': 'prc_', 'showR0': 'Y', 'page': 1},
            headers={
                'User-Agent': 'Mozilla/5.0 (Linux; Android 13; SM-S918N) AppleWebKit/537.36',
                'Accept': 'application/json',
                'Referer': 'https://m.land.naver.com/',
            },
            timeout=10,
        )
        if r.status_code == 200:
            result = r.json().get('result', {})
            if isinstance(result, dict):
                lst = result.get('list', [])
                if lst:
                    auto_name = lst[0].get('atclNm', '')
                    return (name or auto_name), lawd
    except Exception:
        pass

    # 매매 매물 없을 경우 전세/월세로 재시도
    for trade in ('B1', 'B2'):
        try:
            r = requests.get(
                'https://m.land.naver.com/complex/getComplexArticleList',
                params={'hscpNo': no, 'tradTpCd': trade, 'order': 'prc_', 'showR0': 'Y', 'page': 1},
                headers={
                    'User-Agent': 'Mozilla/5.0 (Linux; Android 13; SM-S918N) AppleWebKit/537.36',
                    'Accept': 'application/json',
                    'Referer': 'https://m.land.naver.com/',
                },
                timeout=10,
            )
            if r.status_code == 200:
                result = r.json().get('result', {})
                if isinstance(result, dict):
                    lst = result.get('list', [])
                    if lst:
                        auto_name = lst[0].get('atclNm', '')
                        return (name or auto_name), lawd
        except Exception:
            pass

    # 최후 fallback: fin.land API 시도
    try:
        r = requests.get(
            f'https://fin.land.naver.com/front-api/v1/complex?complexNumber={no}',
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'application/json',
                'Referer': 'https://fin.land.naver.com/',
            },
            timeout=10,
        )
        if r.status_code == 200:
            result = r.json().get('result', {})
            auto_name = result.get('name', '')
            legal_no  = (result.get('address') or {}).get('legalDivisionNumber', '')
            auto_lawd = legal_no[:5] if legal_no else lawd
            return (name or auto_name), (lawd or auto_lawd)
    except Exception:
        pass

    return name or f'단지{no}', lawd

COMPLEX_NO,  _name_arg, _lawd_arg = _ARGS.no, _ARGS.name, _ARGS.lawd
COMPLEX_NAME, LAWD_CD = _resolve_complex_info(COMPLEX_NO, _name_arg, _lawd_arg)

# 파일 이름에 사용할 수 없는 문자 제거
_safe_name = ''.join(c for c in COMPLEX_NAME if c not in r'\/:*?"<>|')

BASE_DIR   = Path(__file__).parent
DB_PATH    = BASE_DIR / f'{_safe_name}.db'
IMAGES_DIR = BASE_DIR / 'images' / COMPLEX_NO
EXCEL_PATH = BASE_DIR / f'{_safe_name}_부동산.xlsm'
_TEMP_XLSX = BASE_DIR / f'{_safe_name}_부동산_tmp.xlsx'
# ──────────────────────────────────────────────────────────────────────────

TODAY       = date.today().strftime('%Y-%m-%d')
TRADE_LABEL = {'A1': '매매', 'B1': '전세', 'B2': '월세'}

BASE_HEADERS = {
    'User-Agent':      'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept-Language': 'ko-KR,ko;q=0.9',
    'Accept':          'application/json, text/plain, */*',
}


_COOKIES_JSON = BASE_DIR / 'naver_cookies.json'


def _get_naver_cookies_from_browser() -> dict:
    """네이버 쿠키 자동 수집 (우선순위 순):
      1) naver_cookies.json 파일 (브라우저 확장 프로그램으로 내보낸 파일)
      2) rookiepy / browser_cookie3 자동 추출 (관리자 권한으로 실행 시 동작)

    [naver_cookies.json 만드는 법]
      Chrome 확장 프로그램 "Cookie Editor" 설치 → land.naver.com 접속(로그인) →
      확장 아이콘 클릭 → Export(JSON) → 파일 이름을 naver_cookies.json 으로 저장
      저장 위치: 이 스크립트와 같은 폴더
    """
    # ── 1) JSON 파일 ─────────────────────────────────────────────────────
    if _COOKIES_JSON.exists():
        try:
            raw = json.loads(_COOKIES_JSON.read_text(encoding='utf-8'))
            # Cookie Editor 형식: [{"name":..., "value":...}, ...]
            if isinstance(raw, list):
                cookies = {c['name']: c['value'] for c in raw
                           if c.get('name') in ('NID_AUT', 'NID_SES')}
            # 단순 dict 형식: {"NID_AUT": ..., "NID_SES": ...}
            elif isinstance(raw, dict):
                cookies = {k: v for k, v in raw.items()
                           if k in ('NID_AUT', 'NID_SES')}
            else:
                cookies = {}
            if len(cookies) >= 2:
                print(f'  쿠키 로드: {_COOKIES_JSON.name} ({len(cookies)}개)')
                return cookies
            else:
                print(f'  경고: {_COOKIES_JSON.name} 에 NID_AUT/NID_SES 없음')
        except Exception as e:
            print(f'  쿠키 JSON 읽기 실패: {e}')

    # ── 2) rookiepy / browser_cookie3 (관리자 권한으로 실행 시 동작) ─────
    for pkg, extractor in _browser_cookie_extractors():
        try:
            cookies = extractor()
            if len(cookies) >= 2:
                return cookies
        except Exception:
            continue

    # ── 3) Playwright 로그인 창 ───────────────────────────────────────────
    return _fetch_cookies_via_playwright()


def _browser_cookie_extractors():
    """(label, callable) 목록 반환 – 설치된 라이브러리에 따라 구성"""
    extractors = []
    try:
        import rookiepy
        for label, fn in [('Chrome', rookiepy.chrome), ('Edge', rookiepy.edge), ('Firefox', rookiepy.firefox)]:
            def _make(f, l):
                def _ext():
                    jar = f(domains=['.naver.com', 'naver.com'])
                    c = {x['name']: x['value'] for x in jar if x['name'] in ('NID_AUT', 'NID_SES')}
                    if len(c) >= 2:
                        print(f'  쿠키 추출 완료 ({l})')
                    return c
                return _ext
            extractors.append((label, _make(fn, label)))
    except ImportError:
        pass
    try:
        import browser_cookie3
        for label, fn in [('Chrome-b', browser_cookie3.chrome), ('Firefox-b', browser_cookie3.firefox)]:
            def _make2(f, l):
                def _ext():
                    jar = f(domain_name='.naver.com')
                    c = {x.name: x.value for x in jar if x.name in ('NID_AUT', 'NID_SES')}
                    if len(c) >= 2:
                        print(f'  쿠키 추출 완료 ({l})')
                    return c
                return _ext
            extractors.append((label, _make2(fn, label)))
    except ImportError:
        pass
    return extractors


def _fetch_cookies_via_playwright() -> dict:
    """Playwright 브라우저 창을 열어 Naver 로그인 후 쿠키 저장"""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print('  playwright 없음: py -3 -m pip install playwright && py -3 -m playwright install chromium')
        return {}

    print('  Naver 로그인 창을 엽니다 (로그인 후 자동으로 쿠키를 저장합니다)...')
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=False,
                args=['--disable-blink-features=AutomationControlled', '--window-size=500,700'],
            )
            ctx  = browser.new_context(
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            )
            page = ctx.new_page()
            page.goto('https://nid.naver.com/nidlogin.login?url=https%3A%2F%2Fland.naver.com%2F', timeout=20000)
            page.bring_to_front()

            # 로그인 완료 = nidlogin 페이지에서 벗어남
            page.wait_for_url(
                lambda url: 'nidlogin' not in url and 'naver.com' in url,
                timeout=120000,
            )
            cookies = {c['name']: c['value'] for c in ctx.cookies()
                       if c['name'] in ('NID_AUT', 'NID_SES')}
            browser.close()

        if len(cookies) >= 2:
            _COOKIES_JSON.write_text(
                json.dumps([{'name': k, 'value': v} for k, v in cookies.items()],
                           ensure_ascii=False, indent=2),
                encoding='utf-8',
            )
            print(f'  쿠키 저장 완료: {_COOKIES_JSON.name}')
            return cookies
        else:
            print('  로그인 후에도 쿠키 없음')
            return {}
    except Exception as e:
        print(f'  Playwright 쿠키 수집 실패: {e}')
        return {}

# ── 컬럼 정의 (A~S = 19열, 최초등록일=R(18), 최종확인일=S(19)) ──────────
LISTING_COLS = [
    '매물ID', '거래유형', '동', '층정보', '공급면적(㎡)', '전용면적(㎡)',
    '호가', '호가(만원)', '방향', '방수', '욕실수',
    '설명', '상세설명', '태그', '중개사', '중개사확인일',
    '이미지수', '최초등록일', '최종확인일',
]

TRADE_COLS = [
    '거래일', '전용면적(㎡)', '층', '거래금액(만원)', '거래금액',
    '거래유형', '동', '매수자', '매도자', '수집일',
]

SUMMARY_COLS = [
    '날짜',
    '매매_건수', '매매_최저가(만원)', '매매_최고가(만원)', '매매_평균가(만원)',
    '전세_건수', '전세_최저가(만원)', '전세_최고가(만원)', '전세_평균가(만원)',
    '월세_건수',
]


# ══════════════════════════════════════════════════════════════════════════
# DB
# ══════════════════════════════════════════════════════════════════════════

_DB_SCHEMA = """
PRAGMA journal_mode=WAL;

CREATE TABLE IF NOT EXISTS listings (
    atcl_no          TEXT PRIMARY KEY,
    trade_type       TEXT,
    building_name    TEXT,
    floor_info       TEXT,
    supply_area      TEXT,
    exclusive_area   TEXT,
    price_display    TEXT,
    price_man        INTEGER,
    direction        TEXT,
    description      TEXT,
    detail_desc      TEXT,
    tags             TEXT,
    realtor_name     TEXT,
    confirm_date     TEXT,
    first_seen_date  TEXT,
    last_seen_date   TEXT,
    article_title    TEXT,
    room_count       TEXT,
    bathroom_count   TEXT,
    total_floor      TEXT,
    build_year       TEXT,
    road_addr        TEXT,
    image_count      INTEGER DEFAULT 0,
    raw_json         TEXT,
    created_at       TEXT DEFAULT (date('now','localtime')),
    updated_at       TEXT DEFAULT (date('now','localtime'))
);

CREATE TABLE IF NOT EXISTS listing_images (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    atcl_no     TEXT NOT NULL,
    image_url   TEXT NOT NULL,
    image_path  TEXT,
    image_type  TEXT,
    seq         INTEGER DEFAULT 0,
    created_at  TEXT DEFAULT (date('now','localtime')),
    UNIQUE(atcl_no, image_url),
    FOREIGN KEY (atcl_no) REFERENCES listings(atcl_no)
);

CREATE TABLE IF NOT EXISTS trades (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    deal_date        TEXT,
    exclusive_area   TEXT,
    floor            TEXT,
    deal_amount_man  INTEGER,
    deal_type        TEXT,
    building_name    TEXT,
    buyer_type       TEXT,
    seller_type      TEXT,
    collected_date   TEXT,
    raw_json         TEXT,
    UNIQUE(deal_date, exclusive_area, floor, deal_amount_man)
);

CREATE TABLE IF NOT EXISTS daily_summary (
    check_date         TEXT PRIMARY KEY,
    sale_count         INTEGER DEFAULT 0,
    sale_min           INTEGER,
    sale_max           INTEGER,
    sale_avg           INTEGER,
    rent_count         INTEGER DEFAULT 0,
    rent_min           INTEGER,
    rent_max           INTEGER,
    rent_avg           INTEGER,
    monthly_rent_count INTEGER DEFAULT 0
);

CREATE INDEX IF NOT EXISTS idx_listings_dates ON listings(first_seen_date, last_seen_date);
CREATE INDEX IF NOT EXISTS idx_images_atcl    ON listing_images(atcl_no);
CREATE INDEX IF NOT EXISTS idx_trades_date    ON trades(deal_date);
"""


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    IMAGES_DIR.mkdir(exist_ok=True)
    with get_db() as conn:
        conn.executescript(_DB_SCHEMA)
        # 기존 레코드 소급: confirm_date(=cfmYmd)가 first_seen_date보다 이전이면 교정
        conn.execute("""
            UPDATE listings
            SET first_seen_date = confirm_date
            WHERE confirm_date != ''
              AND length(confirm_date) = 10
              AND confirm_date < first_seen_date
        """)
        conn.commit()


# ══════════════════════════════════════════════════════════════════════════
# 네트워크
# ══════════════════════════════════════════════════════════════════════════

_auto_cookies: dict | None = None   # 프로세스 내 1회만 추출

def make_session(referer='https://new.land.naver.com/') -> requests.Session:
    global _auto_cookies
    session = requests.Session()
    session.headers.update(BASE_HEADERS)
    session.headers['Referer'] = referer

    if NAVER_COOKIES:
        session.cookies.update(NAVER_COOKIES)
    else:
        if _auto_cookies is None:
            _auto_cookies = _get_naver_cookies_from_browser()
        if _auto_cookies:
            session.cookies.update(_auto_cookies)

    try:
        session.get('https://new.land.naver.com/', timeout=10)
    except Exception:
        pass
    return session


def fetch_all_listings(trade_type='A1') -> list[dict]:
    session = make_session('https://m.land.naver.com/')
    all_items, page = [], 1
    while True:
        r = session.get(
            'https://m.land.naver.com/complex/getComplexArticleList',
            params={'hscpNo': COMPLEX_NO, 'tradTpCd': trade_type,
                    'order': 'prc_', 'showR0': 'Y', 'page': page},
            timeout=10,
        )
        r.raise_for_status()
        data = r.json()
        if data is None:
            break
        result = data.get('result', {})
        if not isinstance(result, dict):
            break
        items  = result.get('list', [])
        all_items.extend(items)
        if result.get('moreDataYn', 'N') != 'Y':
            break
        page += 1
        time.sleep(0.3)
    return all_items


def fetch_article_detail(session: requests.Session, atcl_no: str) -> dict:
    """매물 상세 조회 (이미지·방수·상세설명 등). 실패 시 {}"""
    # 1차: new.land.naver.com
    try:
        r = session.get(
            f'https://new.land.naver.com/api/articles/{atcl_no}',
            params={'complexNo': COMPLEX_NO},
            headers={**BASE_HEADERS,
                     'Referer': f'https://new.land.naver.com/complexes/{COMPLEX_NO}'},
            timeout=10,
        )
        if r.status_code == 200:
            data = r.json()
            if data.get('articleDetail'):
                return data
    except Exception:
        pass

    # 2차: m.land.naver.com
    try:
        r = session.get(
            f'https://m.land.naver.com/article/info/{atcl_no}',
            timeout=10,
        )
        if r.status_code == 200:
            result = r.json().get('result', {})
            if result:
                return result
    except Exception:
        pass

    return {}


class _PlaywrightScraper:
    """fin.land.naver.com 신규 API로 매물 목록 배치 수집 (이미지·상세 포함).
    Playwright stealth 모드로 headless 탐지를 우회합니다.
    per-article 호출 없이 한 번의 POST로 전체 매물을 수집합니다.
    """

    _UA = (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/120.0.0.0 Safari/537.36'
    )
    _BASE = 'https://fin.land.naver.com'
    _LIST_PATH = '/front-api/v1/complex/article/list'

    def __init__(self):
        self._pw      = None
        self._browser = None
        self._ctx     = None
        self._page    = None

    def start(self):
        from playwright.sync_api import sync_playwright
        self._pw      = sync_playwright().start()
        self._browser = self._pw.chromium.launch(
            headless=True,
            args=['--disable-blink-features=AutomationControlled'],
        )
        self._ctx = self._browser.new_context(
            user_agent=self._UA,
            locale='ko-KR',
        )
        # headless 탐지 우회 (navigator.webdriver 숨기기)
        self._ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'plugins',   {get: () => [1,2,3,4,5]});
        """)
        self._page = self._ctx.new_page()
        # 불필요한 리소스 차단
        self._page.route(
            '**/*.{woff,woff2,ttf,otf,png,jpg,jpeg,gif,svg,ico}',
            lambda r: r.abort(),
        )
        try:
            self._page.goto(f'{self._BASE}/home', timeout=20000,
                            wait_until='domcontentloaded')
        except Exception:
            pass
        return self

    def get_lawd_cd(self) -> str:
        """fin.land API에서 단지의 법정동코드 앞 5자리(LAWD_CD) 조회. 실패 시 ''"""
        import json
        try:
            result = self._page.evaluate(f"""
                async () => {{
                    const resp = await fetch('/front-api/v1/complex?complexNumber={COMPLEX_NO}', {{
                        credentials: 'include',
                        headers: {{'Accept': 'application/json'}},
                    }});
                    if (!resp.ok) return null;
                    return await resp.json();
                }}
            """)
            if result and result.get('isSuccess'):
                legal_no = (result.get('result', {}).get('address') or {}).get('legalDivisionNumber', '')
                return legal_no[:5] if legal_no else ''
        except Exception:
            pass
        return ''

    def fetch_all(self) -> dict:
        """fin.land 배치 수집 → {articleNumber(str): representativeArticleInfo} 반환.
        tradeTypes=[] 이면 매매/전세/월세 전부 포함.
        pagination은 lastInfo 커서 방식으로 처리.
        """
        import json, time as _t, random as _r
        all_arts: dict = {}
        body = {
            'size': 30,
            'complexNumber': COMPLEX_NO,
            'tradeTypes': [],
            'pyeongTypes': [],
            'dongNumbers': [],
            'userChannelType': 'PC',
            'articleSortType': 'RANKING_DESC',
            'lastInfo': [],
        }
        page_num = 0
        while True:
            try:
                body_str = json.dumps(body)
                result = self._page.evaluate("""
                    async (bodyStr) => {
                        const resp = await fetch('/front-api/v1/complex/article/list', {
                            method: 'POST',
                            credentials: 'include',
                            headers: {
                                'Accept': 'application/json',
                                'Content-Type': 'application/json',
                            },
                            body: bodyStr,
                        });
                        return {s: resp.status, b: await resp.text()};
                    }
                """, body_str)
            except Exception as e:
                print(f'  fin.land evaluate 오류: {e}')
                break

            if result.get('s') != 200:
                print(f'  fin.land API 오류: {result.get("s")} {result.get("b","")[:100]}')
                break

            try:
                data = json.loads(result['b'])
            except Exception:
                break

            resp = data.get('result', {})
            lst  = resp.get('list', [])
            for item in lst:
                rep = item.get('representativeArticleInfo', {})
                no  = str(rep.get('articleNumber', ''))
                if no:
                    all_arts[no] = rep

            page_num += 1
            print(f'  fin.land page{page_num}: {len(lst)}건 → 누적 {len(all_arts)}건')

            if not resp.get('hasNextPage'):
                break
            body['lastInfo'] = resp.get('lastInfo', [])
            body['seed']     = resp.get('seed', '')
            _t.sleep(_r.uniform(*SCRAPE_DELAY_SEC))

        return all_arts

    def stop(self):
        try:
            if self._browser:
                self._browser.close()
            if self._pw:
                self._pw.stop()
        except Exception:
            pass

    def __enter__(self):
        return self.start()

    def __exit__(self, *_):
        self.stop()


def fetch_molit_real_prices(year_month: str) -> list[dict]:
    if not MOLIT_API_KEY:
        return []
    base_url = (
        'https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade'
        f'?serviceKey={MOLIT_API_KEY}'
    )
    items, page = [], 1
    while True:
        r = requests.get(
            base_url,
            params={'LAWD_CD': LAWD_CD, 'DEAL_YMD': year_month,
                    'numOfRows': 100, 'pageNo': page},
            timeout=15,
        )
        r.raise_for_status()
        root       = ET.fromstring(r.text)
        page_items = [{c.tag: c.text for c in item} for item in root.findall('.//item')]
        items.extend(page_items)
        total = int(root.findtext('.//totalCount') or 0)
        if len(items) >= total or not page_items:
            break
        page += 1
        time.sleep(0.2)
    return [i for i in items if '아너스빌' in (i.get('aptNm') or '')]


def download_image(url: str, atcl_no: str, seq: int) -> str | None:
    """이미지 다운로드 → BASE_DIR 기준 상대 경로 반환 (실패 시 None)"""
    ext = url.split('?')[0].rsplit('.', 1)[-1].lower()
    if ext not in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
        ext = 'jpg'
    save_path = IMAGES_DIR / atcl_no / f'{seq:03d}.{ext}'
    if save_path.exists():
        return str(save_path.relative_to(BASE_DIR))
    try:
        r = requests.get(url, headers=BASE_HEADERS, timeout=15)
        if r.status_code == 200 and r.content:
            save_path.parent.mkdir(parents=True, exist_ok=True)
            save_path.write_bytes(r.content)
            return str(save_path.relative_to(BASE_DIR))
    except Exception:
        pass
    return None


# ══════════════════════════════════════════════════════════════════════════
# 데이터 처리 → DB 저장
# ══════════════════════════════════════════════════════════════════════════

def _parse_price_man(item: dict) -> int | None:
    raw = item.get('sameAddrMinPrc') or item.get('prcInfo') or ''
    s = str(raw).replace(' ', '').replace(',', '')
    try:
        if '억' in s:
            parts = s.split('억')
            return int(parts[0]) * 10000 + (int(parts[1]) if parts[1] else 0)
        return int(s) if s else None
    except Exception:
        return None


def _price_to_man(price_str) -> int | None:
    try:
        return int(str(price_str).replace(',', '').strip())
    except Exception:
        return None


def _parse_naver_date(s) -> str:
    """네이버 날짜 문자열 → 'YYYY-MM-DD'. 파싱 실패 시 TODAY 반환.
    지원 형식: '20260401', '2026.04.01', '2026-04-01'
    """
    import re
    if not s:
        return TODAY
    s = str(s).strip()
    m = re.match(r'^(\d{4})(\d{2})(\d{2})$', s)          # 20260401
    if m:
        return f'{m.group(1)}-{m.group(2)}-{m.group(3)}'
    m = re.match(r'^(\d{4})[.\-/](\d{2})[.\-/](\d{2})', s)  # 2026.04.01
    if m:
        return f'{m.group(1)}-{m.group(2)}-{m.group(3)}'
    return TODAY


def _has_detail(conn: sqlite3.Connection, atcl_no: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM listings WHERE atcl_no=? AND image_count > 0",
        (atcl_no,)
    ).fetchone()
    return row is not None


def collect_and_save_listings(conn: sqlite3.Connection) -> dict:
    """매물 수집 → fin.land 배치 상세조회 → DB 저장. {거래유형: 건수} 반환"""
    import random

    counts  = {}
    scraper = None

    # ── fin.land 배치 수집 (이미지·상세 포함) ──────────────────────────────
    global LAWD_CD
    fin_data: dict = {}   # {atcl_no: representativeArticleInfo}
    if FETCH_DETAIL:
        try:
            print('fin.land 배치 수집 시작...')
            scraper  = _PlaywrightScraper().start()
            # LAWD_CD 미설정이면 fin.land에서 자동 추출
            if not LAWD_CD:
                LAWD_CD = scraper.get_lawd_cd()
                if LAWD_CD:
                    print(f'  LAWD_CD 자동추출: {LAWD_CD}')
            fin_data = scraper.fetch_all()
            print(f'fin.land 배치 수집 완료: {len(fin_data)}건')
        except Exception as e:
            print(f'fin.land 배치 수집 실패 (진행 계속): {e}')

    try:
        for trade_code, trade_label in TRADE_LABEL.items():
            items = fetch_all_listings(trade_code)
            counts[trade_label] = len(items)

            for item in items:
                atcl_no = item.get('atclNo', '')
                if not atcl_no:
                    continue

                deposit       = item.get('tradeRentPrice', 0)
                price_str     = item.get('prcInfo', '')
                price_display = (
                    f"{price_str} / 월 {deposit}만원"
                    if trade_label == '월세' and deposit else price_str
                )
                price_man = _parse_price_man(item)

                # ── fin.land 배치 데이터에서 상세 정보 추출 ─────────────────
                fin_art    = fin_data.get(atcl_no, {})
                fin_detail = fin_art.get('articleDetail', {})
                fin_media  = fin_art.get('articleMedia') or fin_art.get('articleMediaDto') or {}

                # 이미지 목록 (fin.land photos 우선, 없으면 m.land 썸네일)
                photos = [
                    {'photoUrl': p.get('imagePath', ''),
                     'photoTypeCode': p.get('imageType', '')}
                    for p in fin_media.get('photos', [])
                    if p.get('imagePath')
                ]
                if not photos:
                    list_thumb = item.get('repImgUrl', '') or item.get('representativeImgUrl', '')
                    if list_thumb:
                        photos = [{'photoUrl': list_thumb, 'photoTypeCode': 'LIST_THUMB'}]

                # 설명: fin.land articleFeatureDescription 우선
                description = (
                    fin_detail.get('articleFeatureDescription')
                    or item.get('atclFetrDesc')
                    or ''
                ).strip()

                # 방향: fin.land 우선
                direction = fin_detail.get('direction') or item.get('direction', '')

                # 층 정보: fin.land 우선
                floor_info = fin_detail.get('floorInfo') or item.get('flrInfo', '')

                # 확인일: fin.land verificationInfo 우선
                fin_confirm = (fin_art.get('verificationInfo') or {}).get('articleConfirmDate', '')
                confirm_date = fin_confirm or item.get('cfmYmd', '')

                conn.execute("""
                    INSERT INTO listings (
                        atcl_no, trade_type, building_name, floor_info,
                        supply_area, exclusive_area, price_display, price_man,
                        direction, description, detail_desc, tags,
                        realtor_name, confirm_date, first_seen_date, last_seen_date,
                        article_title, room_count, bathroom_count,
                        total_floor, build_year, road_addr,
                        image_count, raw_json, updated_at
                    ) VALUES (
                        :atcl_no, :trade_type, :building_name, :floor_info,
                        :supply_area, :exclusive_area, :price_display, :price_man,
                        :direction, :description, :detail_desc, :tags,
                        :realtor_name, :confirm_date, :first_seen_date, :last_seen_date,
                        :article_title, :room_count, :bathroom_count,
                        :total_floor, :build_year, :road_addr,
                        :image_count, :raw_json, date('now','localtime')
                    )
                    ON CONFLICT(atcl_no) DO UPDATE SET
                        last_seen_date = excluded.last_seen_date,
                        price_display  = excluded.price_display,
                        price_man      = excluded.price_man,
                        confirm_date   = excluded.confirm_date,
                        updated_at     = excluded.updated_at,
                        image_count    = CASE WHEN excluded.image_count > 0
                                             THEN excluded.image_count
                                             ELSE listings.image_count END,
                        detail_desc    = CASE WHEN excluded.detail_desc > ''
                                             THEN excluded.detail_desc
                                             ELSE listings.detail_desc END,
                        article_title  = CASE WHEN excluded.article_title > ''
                                             THEN excluded.article_title
                                             ELSE listings.article_title END,
                        room_count     = CASE WHEN excluded.room_count > ''
                                             THEN excluded.room_count
                                             ELSE listings.room_count END,
                        bathroom_count = CASE WHEN excluded.bathroom_count > ''
                                             THEN excluded.bathroom_count
                                             ELSE listings.bathroom_count END,
                        raw_json       = excluded.raw_json
                """, {
                    'atcl_no':        atcl_no,
                    'trade_type':     trade_label,
                    'building_name':  item.get('bildNm', ''),
                    'floor_info':     floor_info,
                    'supply_area':    str(item.get('spc1', '')),
                    'exclusive_area': str(item.get('spc2', '')),
                    'price_display':  price_display,
                    'price_man':      price_man,
                    'direction':      direction,
                    'description':    description,
                    'detail_desc':    '',
                    'tags':           json.dumps(item.get('tagList', []), ensure_ascii=False),
                    'realtor_name':   item.get('rltrNm', ''),
                    'confirm_date':   confirm_date,
                    'first_seen_date': _parse_naver_date(confirm_date),
                    'last_seen_date':  TODAY,
                    'article_title':  fin_art.get('articleName', ''),
                    'room_count':     '',
                    'bathroom_count': '',
                    'total_floor':    (fin_detail.get('floorDetailInfo') or {}).get('totalFloor', ''),
                    'build_year':     '',
                    'road_addr':      '',
                    'image_count':    len(photos),
                    'raw_json':       json.dumps({**item, '_fin': fin_art}, ensure_ascii=False),
                })

                # 이미지 저장
                if DOWNLOAD_IMAGES and photos:
                    for seq, photo in enumerate(photos):
                        purl = photo.get('photoUrl', '')
                        if not purl:
                            continue
                        local_path = download_image(purl, atcl_no, seq)
                        try:
                            conn.execute("""
                                INSERT OR IGNORE INTO listing_images
                                    (atcl_no, image_url, image_path, image_type, seq)
                                VALUES (?, ?, ?, ?, ?)
                            """, (atcl_no, purl, local_path,
                                  photo.get('photoTypeCode', ''), seq))
                        except Exception:
                            pass

            conn.commit()
            time.sleep(0.5)

    finally:
        if scraper:
            scraper.stop()

    return counts


def collect_and_save_trades(conn: sqlite3.Connection) -> int:
    if not LAWD_CD:
        print('  실거래가 조회 건너뜀 (LAWD_CD 미설정 – --lawd 옵션 또는 자동추출 실패)')
        return 0
    now, total = datetime.now(), 0
    for i in range(MOLIT_MONTHS):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1
        try:
            for t in fetch_molit_real_prices(f'{y}{m:02d}'):
                y2 = t.get('dealYear', '')
                mo = str(t.get('dealMonth', '')).zfill(2)
                d  = str(t.get('dealDay',   '')).zfill(2)
                am = _price_to_man(str(t.get('dealAmount', '')).replace(',', '').strip())
                try:
                    conn.execute("""
                        INSERT OR IGNORE INTO trades
                            (deal_date, exclusive_area, floor, deal_amount_man,
                             deal_type, building_name, buyer_type, seller_type,
                             collected_date, raw_json)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                    """, (
                        f'{y2}-{mo}-{d}',
                        t.get('excluUseAr', ''),
                        t.get('floor', ''),
                        am,
                        t.get('dealingGbn', ''),
                        (t.get('aptDong') or '').strip(),
                        t.get('buyerGbn', ''),
                        t.get('slerGbn', ''),
                        TODAY,
                        json.dumps(t, ensure_ascii=False),
                    ))
                    total += 1
                except Exception:
                    pass
            conn.commit()
            time.sleep(0.3)
        except Exception as e:
            print(f'  실거래가 {y}{m:02d} 조회 오류: {e}')
    return total


def save_daily_summary(conn: sqlite3.Connection):
    rows = conn.execute("""
        SELECT trade_type,
               COUNT(*) AS cnt,
               MIN(price_man) AS mn,
               MAX(price_man) AS mx,
               ROUND(AVG(price_man)) AS av
        FROM listings
        WHERE last_seen_date = ? AND price_man IS NOT NULL
        GROUP BY trade_type
    """, (TODAY,)).fetchall()

    stats = {r['trade_type']: dict(r) for r in rows}
    s  = stats.get('매매', {})
    rt = stats.get('전세', {})
    mr = stats.get('월세', {})

    conn.execute("""
        INSERT INTO daily_summary
            (check_date, sale_count, sale_min, sale_max, sale_avg,
             rent_count, rent_min, rent_max, rent_avg, monthly_rent_count)
        VALUES (?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(check_date) DO UPDATE SET
            sale_count         = excluded.sale_count,
            sale_min           = excluded.sale_min,
            sale_max           = excluded.sale_max,
            sale_avg           = excluded.sale_avg,
            rent_count         = excluded.rent_count,
            rent_min           = excluded.rent_min,
            rent_max           = excluded.rent_max,
            rent_avg           = excluded.rent_avg,
            monthly_rent_count = excluded.monthly_rent_count
    """, (
        TODAY,
        s.get('cnt', 0),  s.get('mn'), s.get('mx'), s.get('av'),
        rt.get('cnt', 0), rt.get('mn'), rt.get('mx'), rt.get('av'),
        mr.get('cnt', 0),
    ))
    conn.commit()


# ══════════════════════════════════════════════════════════════════════════
# 엑셀 내보내기
# ══════════════════════════════════════════════════════════════════════════

_HDR_FILL = PatternFill('solid', fgColor='1F4E79')
_HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
_ALT_FILL = PatternFill('solid', fgColor='D6E4F0')
_BORDER   = Border(
    left=Side(style='thin', color='BFBFBF'),  right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),   bottom=Side(style='thin', color='BFBFBF'),
)
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)


def _style_sheet(ws, col_widths: dict):
    for cell in ws[1]:
        cell.fill = _HDR_FILL; cell.font = _HDR_FONT
        cell.alignment = _CENTER; cell.border = _BORDER
    ws.row_dimensions[1].height = 20
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.border = _BORDER; cell.alignment = _LEFT; cell.font = Font(size=9)
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = 'A2'


def _db_to_listing_df(conn: sqlite3.Connection) -> pd.DataFrame:
    rows = conn.execute("""
        SELECT atcl_no, trade_type, building_name, floor_info,
               supply_area, exclusive_area, price_display, price_man,
               direction, room_count, bathroom_count,
               description, detail_desc, tags,
               realtor_name, confirm_date, image_count,
               first_seen_date, last_seen_date
        FROM listings
        ORDER BY last_seen_date DESC, trade_type, price_man
    """).fetchall()

    data = []
    for r in rows:
        try:
            tags_str = ', '.join(json.loads(r['tags'] or '[]'))
        except Exception:
            tags_str = r['tags'] or ''
        data.append({
            '매물ID':       r['atcl_no'],
            '거래유형':     r['trade_type'],
            '동':           r['building_name'],
            '층정보':       r['floor_info'],
            '공급면적(㎡)': r['supply_area'],
            '전용면적(㎡)': r['exclusive_area'],
            '호가':         r['price_display'],
            '호가(만원)':   r['price_man'],
            '방향':         r['direction'],
            '방수':         r['room_count'],
            '욕실수':       r['bathroom_count'],
            '설명':         r['description'],
            '상세설명':     r['detail_desc'],
            '태그':         tags_str,
            '중개사':       r['realtor_name'],
            '중개사확인일': r['confirm_date'],
            '이미지수':     r['image_count'],
            '최초등록일':   r['first_seen_date'],
            '최종확인일':   r['last_seen_date'],
        })
    return pd.DataFrame(data, columns=LISTING_COLS)


def _db_to_trade_df(conn: sqlite3.Connection) -> pd.DataFrame:
    rows = conn.execute("""
        SELECT deal_date, exclusive_area, floor, deal_amount_man,
               deal_type, building_name, buyer_type, seller_type, collected_date
        FROM trades ORDER BY deal_date DESC
    """).fetchall()
    data = []
    for r in rows:
        am = r['deal_amount_man']
        data.append({
            '거래일':         r['deal_date'],
            '전용면적(㎡)':   r['exclusive_area'],
            '층':             r['floor'],
            '거래금액(만원)': am,
            '거래금액':       f"{am:,}만원" if am else '',
            '거래유형':       r['deal_type'],
            '동':             r['building_name'],
            '매수자':         r['buyer_type'],
            '매도자':         r['seller_type'],
            '수집일':         r['collected_date'],
        })
    return pd.DataFrame(data, columns=TRADE_COLS)


def _db_to_summary_df(conn: sqlite3.Connection) -> pd.DataFrame:
    rows = conn.execute("""
        SELECT check_date, sale_count, sale_min, sale_max, sale_avg,
               rent_count, rent_min, rent_max, rent_avg, monthly_rent_count
        FROM daily_summary ORDER BY check_date DESC
    """).fetchall()
    data = [{
        '날짜':              r['check_date'],
        '매매_건수':         r['sale_count'],
        '매매_최저가(만원)': r['sale_min'],
        '매매_최고가(만원)': r['sale_max'],
        '매매_평균가(만원)': r['sale_avg'],
        '전세_건수':         r['rent_count'],
        '전세_최저가(만원)': r['rent_min'],
        '전세_최고가(만원)': r['rent_max'],
        '전세_평균가(만원)': r['rent_avg'],
        '월세_건수':         r['monthly_rent_count'],
    } for r in rows]
    return pd.DataFrame(data, columns=SUMMARY_COLS)


def _setup_date_view_sheet(wb: openpyxl.Workbook):
    ncols    = len(LISTING_COLS)           # 19
    last_col = get_column_letter(ncols)    # S

    ws = wb.create_sheet(title='날짜별조회', index=0)
    TITLE_FILL = PatternFill('solid', fgColor='1F4E79')
    LABEL_FILL = PatternFill('solid', fgColor='2E75B6')
    INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')
    INFO_FILL  = PatternFill('solid', fgColor='E2EFDA')

    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value = f'  {COMPLEX_NAME}  |  날짜별 매물 조회'
    c.font  = Font(bold=True, size=14, color='FFFFFF')
    c.fill  = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    ws['A2'].value = '조회 날짜'
    ws['A2'].font  = Font(bold=True, color='FFFFFF')
    ws['A2'].fill  = LABEL_FILL
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws['B2'].value  = TODAY
    ws['B2'].font   = Font(bold=True, size=12, color='C00000')
    ws['B2'].fill   = INPUT_FILL
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].number_format = 'YYYY-MM-DD'

    ws.merge_cells(f'C2:{last_col}2')
    ws['C2'].value = '← 날짜를 변경하면 자동으로 해당일 매물이 표시됩니다  (예: 2026-04-01)'
    ws['C2'].font  = Font(italic=True, color='595959', size=10)
    ws['C2'].fill  = INFO_FILL
    ws['C2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 22

    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'].value = '날짜를 입력하면 조회 결과가 표시됩니다.'
    ws['A3'].font  = Font(italic=True, color='7F7F7F', size=10)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 18

    for col_idx, col_name in enumerate(LISTING_COLS, start=1):
        c = ws.cell(row=4, column=col_idx)
        c.value = col_name; c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = _CENTER; c.border = _BORDER
    ws.row_dimensions[4].height = 20
    ws.freeze_panes = 'A5'

    col_widths = {
        'A': 14, 'B': 7,  'C': 10, 'D': 8,  'E': 12, 'F': 12,
        'G': 16, 'H': 12, 'I': 7,  'J': 6,  'K': 6,
        'L': 35, 'M': 40, 'N': 30,
        'O': 20, 'P': 12, 'Q': 8,  'R': 12, 'S': 12,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width


def _save_excel(conn: sqlite3.Connection):
    df_listing = _db_to_listing_df(conn)
    df_trade   = _db_to_trade_df(conn)
    df_summary = _db_to_summary_df(conn)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _setup_date_view_sheet(wb)

    ws = wb.create_sheet('매물_일별요약')
    ws.append(list(df_summary.columns))
    for _, row in df_summary.iterrows():
        ws.append([None if pd.isna(v) else v for v in row])
    _style_sheet(ws, {'A':13,'B':9,'C':16,'D':16,'E':16,'F':9,'G':16,'H':16,'I':16,'J':9})

    ws = wb.create_sheet('매물_누적')
    ws.append(LISTING_COLS)
    for _, row in df_listing.iterrows():
        ws.append([None if pd.isna(v) else v for v in row])
    _style_sheet(ws, {
        'A':14,'B':7,'C':10,'D':8,'E':12,'F':12,
        'G':16,'H':12,'I':7,'J':6,'K':6,
        'L':35,'M':40,'N':30,
        'O':20,'P':12,'Q':8,'R':12,'S':12,
    })

    ws = wb.create_sheet('실거래가_누적')
    ws.append(TRADE_COLS)
    for _, row in df_trade.iterrows():
        ws.append([None if pd.isna(v) else v for v in row])
    _style_sheet(ws, {'A':12,'B':12,'C':6,'D':14,'E':14,'F':10,'G':8,'H':8,'I':8,'J':12})

    wb.save(str(_TEMP_XLSX))


# ── VBA ──────────────────────────────────────────────────────────────────
# 매물_누적 열 구조: A~S (19열)
#   R(18) = 최초등록일,  S(19) = 최종확인일

_VBA_MODULE = '''\
' ── 파일 열 때 자동 실행 ─────────────────────────────────
Sub Auto_Open()
    RefreshDateView
End Sub

' ── 날짜별 매물 조회 ──────────────────────────────────────
Sub RefreshDateView()
    Dim wsSrc   As Worksheet
    Dim wsDst   As Worksheet
    Dim dt      As Date
    Dim dtVal   As Variant
    Dim lastR   As Long
    Dim dstR    As Long
    Dim i       As Long
    Dim c       As Integer
    Dim fs      As String
    Dim ls      As String
    Dim lastCol As Integer
    Dim altFill As Long

    lastCol = 19
    altFill = RGB(214, 228, 240)

    Application.ScreenUpdating = False
    Application.EnableEvents   = False
    On Error GoTo ErrExit

    Set wsSrc = ThisWorkbook.Sheets("매물_누적")
    Set wsDst = ThisWorkbook.Sheets("날짜별조회")

    ' B2 날짜 읽기 (숫자/문자/Date 모두 처리)
    dtVal = wsDst.Range("B2").Value
    If IsEmpty(dtVal) Or dtVal = "" Then
        wsDst.Range("A3").Value = "B2 셀에 조회할 날짜를 입력하세요  (예: 2026-04-01)"
        GoTo Done
    End If
    If IsDate(dtVal) Then
        dt = CDate(dtVal)
    ElseIf IsNumeric(dtVal) Then
        dt = CDate(CLng(dtVal))   ' Excel 시리얼 날짜
    Else
        wsDst.Range("A3").Value = "날짜 형식 오류 – 예) 2026-04-01 로 입력하세요"
        GoTo Done
    End If

    ' 이전 결과 지우기 (5행~)
    Dim lastUsed As Long
    lastUsed = wsDst.Cells(wsDst.Rows.Count, 1).End(xlUp).Row
    If lastUsed >= 5 Then
        wsDst.Rows("5:" & lastUsed).ClearContents
        wsDst.Rows("5:" & lastUsed).Interior.ColorIndex = xlNone
        wsDst.Rows("5:" & lastUsed).Borders.LineStyle   = xlNone
    End If

    lastR = wsSrc.Cells(wsSrc.Rows.Count, 1).End(xlUp).Row
    dstR  = 5

    For i = 2 To lastR
        fs = Trim(CStr(wsSrc.Cells(i, 18).Value))   ' R열: 최초등록일
        ls = Trim(CStr(wsSrc.Cells(i, 19).Value))   ' S열: 최종확인일
        If fs <> "" And ls <> "" And IsDate(fs) And IsDate(ls) Then
            If CDate(fs) <= dt And dt <= CDate(ls) Then
                For c = 1 To lastCol
                    wsDst.Cells(dstR, c).Value = wsSrc.Cells(i, c).Value
                Next c
                With wsDst.Range(wsDst.Cells(dstR, 1), wsDst.Cells(dstR, lastCol))
                    .Font.Size         = 9
                    .Borders.LineStyle = xlContinuous
                    .Borders.Color     = RGB(191, 191, 191)
                    If dstR Mod 2 = 0 Then .Interior.Color = altFill
                End With
                dstR = dstR + 1
            End If
        End If
    Next i

    wsDst.Range("A3").Value = Format(dt, "yyyy-mm-dd") & _
        " 기준 공고중인 매물: 총 " & (dstR - 5) & "건"

Done:
    Application.ScreenUpdating = True
    Application.EnableEvents   = True
    Exit Sub
ErrExit:
    Application.ScreenUpdating = True
    Application.EnableEvents   = True
    MsgBox "오류 " & Err.Number & ": " & Err.Description, vbCritical, "RefreshDateView"
End Sub
'''

_VBA_SHEET_EVENT = '''\
Private Sub Worksheet_Change(ByVal Target As Range)
    If Not Intersect(Target, Me.Range("B2")) Is Nothing Then
        Application.EnableEvents = False
        RefreshDateView
        Application.EnableEvents = True
    End If
End Sub
'''


def _enable_vba_access() -> tuple | None:
    import winreg
    for ver in ('16.0', '15.0', '14.0'):
        try:
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                rf'Software\Microsoft\Office\{ver}\Excel\Security',
                0, winreg.KEY_READ | winreg.KEY_WRITE,
            )
            try:
                old_val, _ = winreg.QueryValueEx(key, 'AccessVBOM')
            except FileNotFoundError:
                old_val = None
            winreg.SetValueEx(key, 'AccessVBOM', 0, winreg.REG_DWORD, 1)
            winreg.CloseKey(key)
            print(f'  레지스트리 VBA 접근 활성화 (Office {ver})')
            return (ver, old_val)
        except Exception:
            continue
    print('  경고: 레지스트리 VBA 접근 설정 실패 – Trust Center 수동 설정 필요')
    return None


def _restore_vba_access(info):
    if info is None:
        return
    import winreg
    ver, old_val = info
    try:
        key = winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            rf'Software\Microsoft\Office\{ver}\Excel\Security',
            0, winreg.KEY_WRITE,
        )
        if old_val is None:
            winreg.DeleteValue(key, 'AccessVBOM')
        else:
            winreg.SetValueEx(key, 'AccessVBOM', 0, winreg.REG_DWORD, old_val)
        winreg.CloseKey(key)
    except Exception:
        pass


def _inject_vba():
    import win32com.client
    import pythoncom

    reg_info = _enable_vba_access()
    pythoncom.CoInitialize()
    xl = win32com.client.Dispatch('Excel.Application')
    xl.Visible       = False
    xl.DisplayAlerts = False

    try:
        wb = xl.Workbooks.Open(os.path.abspath(str(_TEMP_XLSX)))
        print(f'  xlsx 열기 완료 ({wb.Sheets.Count}개 시트)')

        # VBProject 접근 가능 여부 확인
        try:
            vbp = wb.VBProject
        except Exception as e:
            raise RuntimeError(
                f'VBProject 접근 불가 ({e})\n'
                '  → Excel > 파일 > 옵션 > 보안 센터 > 매크로 설정\n'
                '    "VBA 프로젝트 개체 모델에 대한 액세스 신뢰" 체크'
            )

        # 기존 DateViewModule 제거
        for i in range(vbp.VBComponents.Count, 0, -1):
            try:
                if vbp.VBComponents.Item(i).Name == 'DateViewModule':
                    vbp.VBComponents.Remove(vbp.VBComponents.Item(i))
                    break
            except Exception:
                pass

        # RefreshDateView + Auto_Open 모듈 추가
        mod = vbp.VBComponents.Add(1)   # vbext_ct_StdModule
        mod.Name = 'DateViewModule'
        mod.CodeModule.AddFromString(_VBA_MODULE)
        print('  DateViewModule 등록 완료')

        # 날짜별조회 시트 CodeName → Worksheet_Change 이벤트 등록
        target_codename = None
        for i in range(1, wb.Sheets.Count + 1):
            try:
                s = wb.Sheets(i)
                if s.Name == '날짜별조회':
                    target_codename = s.CodeName
                    break
            except Exception:
                pass

        if target_codename:
            sc = vbp.VBComponents(target_codename)
            cm = sc.CodeModule
            if cm.CountOfLines > 0:
                cm.DeleteLines(1, cm.CountOfLines)
            cm.AddFromString(_VBA_SHEET_EVENT)
            print(f'  Worksheet_Change 이벤트 등록 완료 (CodeName={target_codename})')
        else:
            print('  경고: 날짜별조회 시트 CodeName 미발견 – 시트 이벤트 생략')

        # "조회" 버튼 추가 (시트 이벤트 외 직접 실행용)
        try:
            dst = wb.Sheets('날짜별조회')
            # 기존 버튼 제거
            for btn in list(dst.Buttons()):
                if btn.Caption == '조회':
                    btn.Delete()
            # B2 셀 오른쪽 (D2 위치 근처) 에 버튼 배치
            btn = dst.Buttons().Add(310, 6, 50, 20)
            btn.Caption   = '조회'
            btn.OnAction  = 'RefreshDateView'
            btn.Font.Size = 10
            btn.Font.Bold = True
            print('  조회 버튼 추가 완료')
        except Exception as e:
            print(f'  조회 버튼 추가 실패 (무시): {e}')

        wb.SaveAs(os.path.abspath(str(EXCEL_PATH)), FileFormat=52)
        wb.Close(SaveChanges=False)
        print(f'  xlsm 저장: {EXCEL_PATH}')
    finally:
        xl.Quit()
        pythoncom.CoUninitialize()
        _restore_vba_access(reg_info)

    if _TEMP_XLSX.exists():
        _TEMP_XLSX.unlink()


def export_to_excel(conn: sqlite3.Connection):
    _save_excel(conn)
    n_listing = conn.execute("SELECT COUNT(*) FROM listings").fetchone()[0]
    n_trade   = conn.execute("SELECT COUNT(*) FROM trades").fetchone()[0]
    n_summary = conn.execute("SELECT COUNT(*) FROM daily_summary").fetchone()[0]
    print(f'  매물_누적: {n_listing}행 / 실거래가_누적: {n_trade}행 / 일별요약: {n_summary}행')

    try:
        _inject_vba()
    except Exception as e:
        print(f'\n  [VBA 실패] {e}')
        print('  ─ 해결 방법 ─────────────────────────────────────────')
        print('  Excel > 파일 > 옵션 > 보안 센터 > 보안 센터 설정')
        print('  → 매크로 설정 탭 > "VBA 프로젝트 개체 모델에 대한 액세스 신뢰" 체크')
        print('  ─────────────────────────────────────────────────────')
        import shutil
        xlsx_path = EXCEL_PATH.with_suffix('.xlsx')
        if _TEMP_XLSX.exists():
            shutil.copy(str(_TEMP_XLSX), str(xlsx_path))
            _TEMP_XLSX.unlink()
        print(f'  (임시) VBA 없이 xlsx로 저장: {xlsx_path}')


# ══════════════════════════════════════════════════════════════════════════
# 콘솔 출력
# ══════════════════════════════════════════════════════════════════════════

def print_summary(conn: sqlite3.Connection):
    print(f"\n{'='*60}")
    print(f"  {COMPLEX_NAME}  |  {TODAY}")
    print(f"{'='*60}")

    rows = conn.execute("""
        SELECT trade_type, COUNT(*) AS cnt,
               MIN(price_man) AS mn, MAX(price_man) AS mx
        FROM listings WHERE last_seen_date = ?
        GROUP BY trade_type
    """, (TODAY,)).fetchall()
    for r in rows:
        mn = f"{r['mn']:,}" if r['mn'] else '-'
        mx = f"{r['mx']:,}" if r['mx'] else '-'
        print(f"  [{r['trade_type']}] {r['cnt']}건  "
              f"최저 {mn}만원 ~ 최고 {mx}만원")

    total_l = conn.execute("SELECT COUNT(*) FROM listings").fetchone()[0]
    total_i = conn.execute("SELECT COUNT(*) FROM listing_images").fetchone()[0]
    total_t = conn.execute("SELECT COUNT(*) FROM trades").fetchone()[0]
    print(f"\n  DB 누적: 매물 {total_l}건 / 이미지 {total_i}장 / 실거래가 {total_t}건")

    trades = conn.execute("""
        SELECT deal_date, exclusive_area, floor, deal_amount_man, deal_type
        FROM trades ORDER BY deal_date DESC LIMIT 5
    """).fetchall()
    if trades:
        print(f"\n  최근 실거래가")
        print(f"  {'거래일':<12} {'전용면적':>8}  {'층':>4}  {'거래금액':>12}  거래유형")
        print(f"  {'-'*52}")
        for t in trades:
            am = f"{t['deal_amount_man']:,}만원" if t['deal_amount_man'] else '-'
            print(f"  {t['deal_date']:<12} {str(t['exclusive_area']):>7}㎡  "
                  f"{str(t['floor']):>3}층  {am:>12}  {t['deal_type']}")


# ══════════════════════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print(f"[{TODAY}] 데이터 수집 시작...")

    init_db()

    with get_db() as conn:
        print("  매물 조회 중...")
        counts = collect_and_save_listings(conn)
        for label, cnt in counts.items():
            print(f"    [{label}] {cnt}건")

        print("  실거래가 조회 중...")
        trade_cnt = collect_and_save_trades(conn)
        print(f"  → {trade_cnt}건 (신규)")

        save_daily_summary(conn)
        print_summary(conn)

        print("\n  엑셀 내보내기...")
        export_to_excel(conn)
